import os
import json
import logging
import httpx
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Any
from collections import defaultdict
import openpyxl

from config import ClientConfig, settings

logger = logging.getLogger(__name__)

class ReportEngine:
    def __init__(self):
        self.client = httpx.AsyncClient(timeout=30.0)

    async def _read_logs_for_days(self, client_id: str, days: int) -> List[Dict[str, Any]]:
        """Reads JSONL out of logs/{client_id}/leads/ for the last N days."""
        all_leads = []
        now = datetime.now(timezone.utc)
        
        log_dir = os.path.join(settings.log_dir, client_id, "leads")
        if not os.path.exists(log_dir):
            return all_leads

        for i in range(days):
            target_date = (now - timedelta(days=i)).strftime("%Y-%m-%d")
            filepath = os.path.join(log_dir, f"{target_date}.jsonl")
            
            if os.path.exists(filepath):
                with open(filepath, "r", encoding="utf-8") as f:
                    for line in f:
                        if not line.strip():
                            continue
                        try:
                            data = json.loads(line)
                            all_leads.append(data)
                        except json.JSONDecodeError:
                            logger.error(f"[{client_id}] Malformed JSON line in {filepath}")
                            
        return all_leads

    async def send_daily_digest(self, client_id: str, config: ClientConfig):
        """Phase 2 placeholder removed, Phase 4 implemented via dispatch but daily digest trigger isn't required by user reqs."""
        logger.info(f"[{client_id}] Daily digest triggered")
        leads = await self._read_logs_for_days(client_id, 1)
        
        # Build payload
        payload = {
            "type": "daily_digest",
            "client_id": client_id,
            "business_name": config.business_name,
            "total_leads_today": len(leads),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }

        if config.n8n_webhook_url:
            try:
                await self.client.post(config.n8n_webhook_url, json=payload)
                logger.info(f"[{client_id}] Dispatched daily digest to n8n")
            except Exception as e:
                logger.error(f"[{client_id}] Failed to dispatch daily digest: {e}")

    async def send_weekly_report(self, client_id: str, config: ClientConfig):
        """Generates the weekly Excel and dispatches the payload to n8n."""
        logger.info(f"[{client_id}] Weekly report triggered")
        leads = await self._read_logs_for_days(client_id, 7)
        
        if not leads:
            logger.info(f"[{client_id}] No leads found for the weekly report.")
            return

        excel_path = self._generate_excel(client_id, leads)
        
        # We need to compile the window.REPORT_DATA equivalent to send to n8n
        payload = {
            "type": "weekly_report",
            "client_id": client_id,
            "business_name": config.business_name,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "metrics": {
                "total_chats": sum(len(lead.get("messages", [])) // 2 for lead in leads) if sum(len(lead.get("messages", [])) // 2 for lead in leads) > 0 else sum(1 for lead in leads),
                "leads_captured": sum(1 for lead in leads if lead.get("user_name") and lead.get("user_phone")),
                "avg_confidence": sum(l.get("confidence", 0) for l in leads) / len(leads) if leads else 0.0
            }
        }

        # Dispatch
        if config.n8n_webhook_url:
            try:
                # To send an excel file, normally we'd do multipart/form-data. For simplicity, we just trigger n8n,
                # n8n can generate the email body. Since we generate the excel file locally here, n8n either needs access 
                # to the filesystem (if self-hosted) or we send the base64 encoded excel.
                import base64
                with open(excel_path, "rb") as xf:
                    encoded_excel = base64.b64encode(xf.read()).decode("utf-8")
                    payload["excel_attachment"] = {
                        "filename": os.path.basename(excel_path),
                        "content": encoded_excel,
                        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    }
                
                await self.client.post(config.n8n_webhook_url, json=payload)
                logger.info(f"[{client_id}] Dispatched weekly report to n8n")
            except Exception as e:
                logger.error(f"[{client_id}] Failed to dispatch weekly report: {e}")

    def _generate_excel(self, client_id: str, leads: List[Dict[str, Any]]) -> str:
        wb = openpyxl.Workbook()
        
        # Sheet 1: Lead Summary
        ws1 = wb.active
        ws1.title = "Lead Summary"
        ws1.append(["Date", "Name", "Phone", "Intent", "Requirement", "Budget", "Confidence", "Status"])
        for lead in leads:
            ws1.append([
                lead.get("timestamp", ""),
                lead.get("user_name", ""),
                lead.get("user_phone", ""),
                lead.get("intent", ""),
                lead.get("requirement", ""),
                lead.get("budget", ""),
                lead.get("confidence", 0.0),
                "Captured" if (lead.get("user_name") and lead.get("user_phone")) else "Incomplete"
            ])

        # Sheet 2: Conversation Analytics (Daily stats)
        ws2 = wb.create_sheet(title="Conversation Analytics")
        ws2.append(["Date", "Total Chats", "Leads Captured", "Avg Confidence"])
        daily_stats = defaultdict(lambda: {"total": 0, "total_msgs": 0, "leads": 0, "conf_sum": 0.0})
        for lead in leads:
            date_str = lead.get("timestamp", "")[:10]
            daily_stats[date_str]["total"] += 1
            daily_stats[date_str]["total_msgs"] += len(lead.get("messages", [])) // 2
            if lead.get("user_name") and lead.get("user_phone"):
                daily_stats[date_str]["leads"] += 1
            daily_stats[date_str]["conf_sum"] += lead.get("confidence", 0.0)
            
        for date_str, stats in sorted(daily_stats.items()):
            avg_conf = stats["conf_sum"] / stats["total"] if stats["total"] else 0.0
            chats_count = stats["total_msgs"] if stats["total_msgs"] > 0 else stats["total"]
            ws2.append([date_str, chats_count, stats["leads"], round(avg_conf, 2)])

        # Sheet 3: Intent & Topic Breakdown
        ws3 = wb.create_sheet(title="Intent & Topic Breakdown")
        ws3.append(["Topic", "Times Asked", "Lead Conversion Rate", "Avg Confidence"])
        intent_stats = defaultdict(lambda: {"count": 0, "leads": 0, "conf_sum": 0.0})
        for lead in leads:
            intent = lead.get("intent", "unknown")
            intent_stats[intent]["count"] += 1
            if lead.get("user_name") and lead.get("user_phone"):
                intent_stats[intent]["leads"] += 1
            intent_stats[intent]["conf_sum"] += lead.get("confidence", 0.0)
            
        for intent, stats in sorted(intent_stats.items()):
            conv_rate = (stats["leads"] / stats["count"]) * 100 if stats["count"] else 0.0
            avg_conf = stats["conf_sum"] / stats["count"] if stats["count"] else 0.0
            ws3.append([intent, stats["count"], f"{conv_rate:.1f}%", round(avg_conf, 2)])

        reports_dir = os.path.join(settings.log_dir, client_id, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        report_path = os.path.join(reports_dir, f"weekly_report_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx")
        wb.save(report_path)
        return report_path

    async def close(self):
        await self.client.aclose()
